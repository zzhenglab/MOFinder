"""
Excel I/O shared by Step 2.1 / 2.2.

The two pipelines write progress into slightly different columns:

  - Step 2.1 writes ``Downloaded``      ('1' = ok, '0' = failed, '' = pending)
  - Step 2.2 writes ``SI Downloaded``   (same value set)

so the load / save helpers below take the status column name as a
parameter. Both workbooks also get a ``DOI Link`` column with a clickable
hyperlink applied via openpyxl.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from .doi import doi_to_link


# ===========================================================================
# Hyperlinks
# ===========================================================================
def apply_hyperlinks(xlsx_path: Path, log=print) -> None:
    """
    Apply hyperlink styling to every non-empty value in the ``DOI Link`` column.

    Reads the workbook back from disk, scans the header row case-insensitively
    for "doi link", and writes ``Hyperlink`` style + URL on each populated
    cell below.
    """
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if isinstance(v, str) and v.strip().lower() == "doi link":
                col = c
                break
        if col is None:
            wb.save(xlsx_path)
            return
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            url = cell.value
            if url and isinstance(url, str) and url.strip():
                cell.hyperlink = url
                cell.style = "Hyperlink"
        wb.save(xlsx_path)
    except Exception as e:
        log("Hyperlink apply error:", e)


# ===========================================================================
# Status-column value normalization
# ===========================================================================
def status_norm(val) -> str:
    """
    Map a free-form status cell to one of ``'1'`` / ``'0'`` / ``''``.

    Accepts ``1 / 1.0 / true / yes / y`` -> ``'1'``,
            ``0 / 0.0 / false / no / n`` -> ``'0'``, else ``''``.
    """
    s = str(val).strip().lower()
    if s in ("1", "1.0", "true", "yes", "y"):
        return "1"
    if s in ("0", "0.0", "false", "no", "n"):
        return "0"
    return ""


# ===========================================================================
# Load + prepare the input workbook
# ===========================================================================
def load_and_prepare_excel(
    path: Path,
    *,
    status_column: str,
    write_back: bool = True,
    apply_links: bool = True,
    normalize_status: bool = False,
    log=print,
) -> pd.DataFrame:
    """
    Read the input workbook, validate columns, add ``DOI Link``, ensure the
    status column exists, and (optionally) write the prepared file back to
    disk with hyperlinks applied.

    Raises ``ValueError`` if ``DOI`` or ``Publisher`` is missing.

    ``normalize_status=True`` runs ``status_norm`` over the status column
    after loading (Step 2.2's behavior).
    """
    log("Loading Excel:", path)
    df = pd.read_excel(path)
    if "DOI" not in df.columns:
        raise ValueError("Missing column: DOI")
    if "Publisher" not in df.columns:
        raise ValueError("Missing column: Publisher")

    df["DOI"] = df["DOI"].astype(str)
    df["DOI Link"] = df["DOI"].apply(doi_to_link)
    if status_column not in df.columns:
        df[status_column] = ""
    if normalize_status:
        df[status_column] = df[status_column].apply(status_norm)

    if write_back:
        df.to_excel(path, index=False)
        if apply_links:
            apply_hyperlinks(path, log=log)

    log("Excel prepared. Rows:", len(df))
    return df


# ===========================================================================
# Save the in-memory df back to disk
# ===========================================================================
def save_progress(
    df: pd.DataFrame,
    path: Path,
    *,
    apply_links: bool = False,
    log=print,
) -> None:
    """Write ``df`` to ``path``, optionally re-applying ``DOI Link`` hyperlinks."""
    df.to_excel(path, index=False)
    if apply_links:
        apply_hyperlinks(path, log=log)
    log("Saved workbook:", path)


# ===========================================================================
# Download folder
# ===========================================================================
def ensure_download_dir(excel_path: Path, name: str = "downloaded") -> Path:
    """
    Make sure the per-Excel download folder exists; return its Path.

    The folder lives next to the Excel file, named ``name``
    (``"downloaded"`` for Step 2.1, ``"SI downloaded"`` for Step 2.2).
    """
    out = excel_path.parent / name
    out.mkdir(exist_ok=True)
    return out
